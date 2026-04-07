import csv
import os
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def csv_to_xlsx(csv_path: str, xlsx_path: str) -> None:
    """
    Конвертирует CSV в XLSX.

    Args:
        csv_path: Путь к исходному CSV файлу
        xlsx_path: Путь для сохранения XLSX файла

    Raises:
        FileNotFoundError: Если файл не существует
        ValueError: Если файл не CSV, пустой, невалидный CSV или не UTF-8
    """
    # Проверка существования файла
    if not os.path.exists(csv_path):
        raise FileNotFoundError(f"Файл {csv_path} не найден")

    # Проверка расширения файла
    if not csv_path.lower().endswith(".csv"):
        raise ValueError(f"Файл {csv_path} не является CSV файлом")

    data = []

    try:
        with open(csv_path, "r", encoding="utf-8") as csv_file:
            # Проверка что файл не пустой
            content = csv_file.read().strip()
            if not content:
                raise ValueError(f"Файл {csv_path} пуст")

            csv_file.seek(0)  # Возвращаемся к началу файла
            reader = csv.reader(csv_file)  # создается обект для чтения построчнл

            for row in reader:
                data.append(row)

    except UnicodeDecodeError:
        raise ValueError(f"Файл {csv_path} имеет неверную кодировку (требуется UTF-8)")
    except csv.Error as e:
        raise ValueError(f"Невалидный CSV в файле {csv_path}: {e}")

    if not data:
        raise ValueError(f"CSV файл {csv_path} не содержит данных")

    # Создаем новую книгу Excel
    wb = Workbook()  # создает новую книгу эксель
    ws = wb.active  # получает активный лист
    ws.title = "Sheet1"  # название листа

    # Записываем данные в лист
    for row_idx, row_data in enumerate(
        data, 1
    ):  # idx-номер строки,  enumerate- цикл по значениям строки с индексом 1
        for col_idx, cell_value in enumerate(
            row_data, 1
        ):  # записывает значение в ячейку
            ws.cell(row=row_idx, column=col_idx, value=cell_value)  # ячейка в таблице

    # Настраиваем автоширину колонок (не менее 8 символов)
    for col_idx in range(1, len(data[0]) + 1):
        column_letter = get_column_letter(col_idx)  # преобразует номер в букву колонки
        max_length = 8  # минимальная ширина строки

        for row_idx in range(1, len(data) + 1):
            cell_value = ws.cell(
                row=row_idx, column=col_idx
            ).value  # получает значение ячейки
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))

        ws.column_dimensions[column_letter].width = (
            max_length + 2
        )  # добавляем небольшой отступ

    # Сохраняем файл
    wb.save(xlsx_path)


if __name__ == "__main__":

    test_cases = [
        # (аргументы, ожидаемая_ошибка, описание)
        (("data/people.csv", "data/people.xlsx"), None, "НОРМАЛЬНЫЙ CSV"),
        (("data/peopleempty.csv", "data/people2.xlsx"), ValueError, "ПУСТОЙ ФАЙЛ"),
        (
            ("data/peoplenotexcist.csv", "data/people2.xlsx"),
            FileNotFoundError,
            "ФАЙЛ НЕ СУЩЕСТВУЕТ",
        ),
        (
            ("data/people1251.csv", "data/people2.xlsx"),
            ValueError,
            "НЕ UTF-8 КОДИРОВКА",
        ),
        (("data/people.txt", "data/people2.xlsx"), ValueError, "НЕ CSV РАСШИРЕНИЕ"),
    ]

    for args, expected_error, description in test_cases:
        csv_path, xlsx_path = args  # распаковка аргумента
        print(f"{description}")
        print(f"{csv_path}")

        try:
            csv_to_xlsx(csv_path, xlsx_path)
            if expected_error is None:
                print("   УСПЕШНО - файл создан")
                if os.path.exists(xlsx_path):
                    print("    XLSX файл существует на диске")
            else:
                print(
                    f"    ОШИБКА: ожидалась {expected_error.__name__}, но функция завершилась успешно"
                )

        except Exception as e:
            if expected_error is not None and isinstance(e, expected_error):
                print(f"   КОРРЕКТНАЯ ОШИБКА: {e}")
            else:
                print(f"   НЕОЖИДАННАЯ ОШИБКА: {e} (ожидался {expected_error})")

    # Дополнительная проверка созданных файлов
    success_files = []
    for file in os.listdir("data"):  # получает список файлов в папке дата
        if file.endswith(".xlsx"):
            success_files.append(file)
