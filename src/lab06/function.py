import csv
import json
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import os


def json_to_csv(json_path: str, csv_path: str) -> None:
    # Проверка расширения файла
    if not json_path.lower().endswith(".json"):
        raise ValueError(f"Файл {json_path} не является JSON файлом")

    # Проверка существования файла
    if not os.path.exists(json_path):
        raise FileNotFoundError(f"JSON файл не найден: {json_path}")

    try:
        with open(json_path, "r", encoding="UTF-8") as json_file:
            content = json_file.read().strip()

            # Проверка на пустой файл
            if not content:
                raise ValueError(f"JSON файл пуст: {json_path}")

            data = json.loads(content)
            # преобразет джисон строку в питон объект,список или словарь

    except UnicodeDecodeError:
        raise ValueError(f"Файл {json_path} имеет неверную кодировку (требуется UTF-8)")
    except json.JSONDecodeError as e:  # сохраняет объект исключения в переменную е
        raise ValueError(f"Невалидный JSON в файле {json_path}: {e}")

    # Проверка структуры данных
    if not isinstance(data, list):
        raise ValueError("JSON должен содержать список")

    if not data:
        raise ValueError("JSON должен содержать непустой список")

    for item in data:  # проходимся по элементам списка
        if not isinstance(item, dict):
            raise ValueError("Все элементы в JSON должны быть словарями")

    # Формирование и запись CSV
    all_columns = []
    for item in data:
        for key in item.keys():  # возвращает все ключи словаря
            if key not in all_columns:  # проверяет есть ли такой ключ в списке
                all_columns.append(key)

    with open(csv_path, "w", encoding="UTF-8", newline="") as csv_file:
        writer = csv.DictWriter(
            csv_file, fieldnames=all_columns
        )  # создает writer для записи словарей в csv
        # fieildnames=передает список заголовков столбцов
        writer.writeheader()
        for row in data:
            full_row = {}
            for column in all_columns:
                full_row[column] = row.get(
                    column, ""
                )  # получает значение из словаря или пустую строку если ключа нет
            writer.writerow(full_row)


def csv_to_json(csv_path: str, json_path: str) -> None:
    # Проверка расширения файла
    if not csv_path.lower().endswith(".csv"):
        raise ValueError(f"Файл {csv_path} не является CSV файлом")

    # Проверка существования файла
    if not os.path.exists(csv_path):
        raise FileNotFoundError(f"CSV файл не найден: {csv_path}")

    try:
        with open(csv_path, "r", encoding="UTF-8") as csv_file:
            content = csv_file.read().strip()

            # Проверка на пустой файл
            if not content:
                raise ValueError(f"CSV файл пуст: {csv_path}")

            csv_file.seek(0)
            reader = csv.DictReader(csv_file)

            # Проверка наличия заголовков
            if reader.fieldnames is None:
                raise ValueError("CSV файл не содержит заголовка")

            data = []
            for row in reader:
                str_row = {}
                for key, value in row.items():
                    str_row[key] = str(value)
                data.append(str_row)

            if not data:
                raise ValueError("CSV файл не содержит данных")

    except UnicodeDecodeError:
        raise ValueError(f"Файл {csv_path} имеет неверную кодировку (требуется UTF-8)")
    except csv.Error as e:
        raise ValueError(f"Невалидный CSV в файле {csv_path}: {e}")

    with open(json_path, "w", encoding="UTF-8") as json_file:
        json.dump(data, json_file, ensure_ascii=False, indent=2)


def csv_to_xlsx(csv_path: str, xlsx_path: str) -> None:
    """
    Конвертирует CSV в XLSX.

    Args:
        csv_path: Путь к исходному CSV файлу
        xlsx_path: Путь для сохранения XLSX файла

    Raises:
        FileNotFoundError: Если файл не существует
        ValueError: Если файл не CSV, пустой, невалидный CSV или не UTF-8
    """
    # Проверка существования файла
    if not os.path.exists(csv_path):
        raise FileNotFoundError(f"Файл {csv_path} не найден")

    # Проверка расширения файла
    if not csv_path.lower().endswith(".csv"):
        raise ValueError(f"Файл {csv_path} не является CSV файлом")

    data = []

    try:
        with open(csv_path, "r", encoding="utf-8") as csv_file:
            # Проверка что файл не пустой
            content = csv_file.read().strip()
            if not content:
                raise ValueError(f"Файл {csv_path} пуст")

            csv_file.seek(0)  # Возвращаемся к началу файла
            reader = csv.reader(csv_file)  # создается обект для чтения построчнл

            for row in reader:
                data.append(row)

    except UnicodeDecodeError:
        raise ValueError(f"Файл {csv_path} имеет неверную кодировку (требуется UTF-8)")
    except csv.Error as e:
        raise ValueError(f"Невалидный CSV в файле {csv_path}: {e}")

    if not data:
        raise ValueError(f"CSV файл {csv_path} не содержит данных")

    # Создаем новую книгу Excel
    wb = Workbook()  # создает новую книгу эксель
    ws = wb.active  # получает активный лист
    ws.title = "Sheet1"  # название листа

    # Записываем данные в лист
    for row_idx, row_data in enumerate(
        data, 1
    ):  # idx-номер строки,  enumerate- цикл по значениям строки с индексом 1
        for col_idx, cell_value in enumerate(
            row_data, 1
        ):  # записывает значение в ячейку
            ws.cell(row=row_idx, column=col_idx, value=cell_value)  # ячейка в таблице

    # Настраиваем автоширину колонок (не менее 8 символов)
    for col_idx in range(1, len(data[0]) + 1):
        column_letter = get_column_letter(col_idx)  # преобразует номер в букву колонки
        max_length = 8  # минимальная ширина строки

        for row_idx in range(1, len(data) + 1):
            cell_value = ws.cell(
                row=row_idx, column=col_idx
            ).value  # получает значение ячейки
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))

        ws.column_dimensions[column_letter].width = (
            max_length + 2
        )  # добавляем небольшой отступ

    # Сохраняем файл
    wb.save(xlsx_path)

def top_n(args):
    """
    Обрабатывает команду stats - статистика по словам в файле
    """
    try:
        # Читаем файл
        with open(args.input, 'r', encoding='utf-8') as file:
            text = file.read()
        
        # Разбиваем на слова и подсчитываем частоту
        words = text.lower().split()
        freq = {}
        for word in words:
            # Убираем знаки препинания
            word = word.strip('.,!?;:"()[]')
            if word:  # если слово не пустое
                freq[word] = freq.get(word, 0) + 1
        
        # Сортируем по частоте
        items = []
        for word, count in freq.items():
            items.append([word, count])
        items.sort(key=lambda x: x[1], reverse=True)
        
        # Выводим топ-N слов
        n = args.top if hasattr(args, 'top') else 10
        top_words = items[:n]
        
        print(f"Топ-{n} самых частых слов:")
        print("-" * 30)
        for i, (word, count) in enumerate(top_words, 1):
            print(f"{i:2}. {word:15} - {count:3} раз")
            
        return top_words
        
    except Exception as e:
        print(f"Ошибка при обработке файла: {e}")
        return []
